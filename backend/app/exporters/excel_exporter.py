"""Generate Excel workbooks from session results."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _coerce_step_text(item: Any) -> str:
    """Convert a step item (str/dict/other) into readable text.

    Some LLM outputs may structure steps as dict objects, e.g. {"step": "..."}.
    This helper extracts common fields or falls back to string casting.
    """
    if item is None:
        return ""
    if isinstance(item, str):
        return item.strip()
    if isinstance(item, dict):
        for key in ("text", "desc", "description", "action", "step", "content", "title"):
            value = item.get(key)
            if isinstance(value, str):
                return value.strip()
        try:
            return " ".join(str(v).strip() for v in item.values() if v is not None)
        except Exception:
            return str(item)
    return str(item).strip()


def _format_steps(steps: Any) -> str:
    """Normalize various step formats into a newline-joined string with numbering."""
    if steps is None:
        return ""
    if isinstance(steps, (list, tuple, set)):
        steps_list = list(steps)
        if len(steps_list) == 1:
            return _coerce_step_text(steps_list[0])
        return "\n".join(f"{i+1}. {_coerce_step_text(x)}" for i, x in enumerate(steps_list))
    return _coerce_step_text(steps)


def _format_field_with_numbering(field: Any) -> str:
    """Format field (preconditions/expected) with numbering if it's a list."""
    if field is None:
        return ""
    if isinstance(field, (list, tuple)):
        field_list = list(field)
        if len(field_list) == 0:
            return ""
        if len(field_list) == 1:
            return _coerce_step_text(field_list[0])
        return "\n".join(f"{i+1}. {_coerce_step_text(x)}" for i, x in enumerate(field_list))
    return _coerce_step_text(field)


def generate_excel_bytes(test_cases: Dict[str, Any]) -> bytes:
    """Create an Excel workbook from structured test cases."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "测试用例汇总"

    headers = ["用例ID", "模块", "用例标题", "前置条件", "操作步骤", "预期结果", "优先级"]
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        # Use 8-digit ARGB to avoid color parsing issues in some openpyxl versions
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for module in test_cases.get("modules", []):
        module_name = module.get("name", "模块")
        for case in module.get("cases", []):
            worksheet.append(
                [
                    case.get("id", ""),
                    module_name,
                    case.get("title", ""),
                    _format_field_with_numbering(case.get("preconditions", "")),
                    _format_steps(case.get("steps", [])),
                    _format_field_with_numbering(case.get("expected", "")),
                    str(case.get("priority", "P2") or "P2"),
                ]
            )

    for column_cells in worksheet.columns:
        lengths = [len(str(cell.value or "")) for cell in column_cells]
        width = min(60, max(lengths, default=10) + 4)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width

    stream = BytesIO()
    workbook.save(stream)
    buffer = stream.getvalue()
    stream.close()
    return buffer
